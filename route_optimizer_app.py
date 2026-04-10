"""
Route Optimizer — Streamlit App  v7
Combines routing_osv_ver5 (distance fetching) + redistribution_dist_ver6 (optimization)

Changes vs v6:
- sh1: added `address` (delivery point address) and `zero_address` (driver home address)
- sh2: added `OwnerName` — which driver uses the track on a given day
- sh3: added `odometr` — odometer reading for the last period of a track
- Output Маршрути: СТАРТ/ФІНІШ rows now show zero_address instead of "база"
- Output Зведення: new Одометр column, calculated backwards from the last period
- Multi-owner support: one track can be shared by multiple OwnerName (not on the same day)
"""

import math
import io
import os
import time
import random
import sqlite3
import threading
import tempfile
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ═══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Route Optimizer",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═══════════════════════════════════════════════════════════════════
#  CUSTOM CSS
# ═══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
    .block-container { max-width: 1200px; }

    .main-title {
        font-weight: 800;
        font-size: 2.2rem;
        background: linear-gradient(135deg, #1a73e8 0%, #00bcd4 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.2rem;
    }
    .sub-title {
        color: #666;
        font-size: 0.95rem;
        margin-bottom: 1.5rem;
    }

    .stat-card {
        background: linear-gradient(135deg, #f8f9ff 0%, #e8f4fd 100%);
        border: 1px solid #d0e4f5;
        border-radius: 12px;
        padding: 1rem 1.2rem;
        text-align: center;
    }
    .stat-card h3 {
        font-family: monospace;
        font-size: 1.5rem;
        color: #1a73e8;
        margin: 0;
    }
    .stat-card p {
        color: #555;
        font-size: 0.8rem;
        margin: 0.2rem 0 0 0;
    }

    div[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0d1b2a 0%, #1b2838 100%);
    }
    div[data-testid="stSidebar"] * {
        color: #e0e0e0 !important;
    }
    div[data-testid="stSidebar"] .stSlider > div > div > div {
        color: #4fc3f7 !important;
    }

    .log-box {
        font-family: monospace;
        font-size: 0.75rem;
        background: #0d1117;
        color: #58a6ff;
        padding: 1rem;
        border-radius: 10px;
        max-height: 400px;
        overflow-y: auto;
        line-height: 1.6;
    }

    .coord-info {
        background: #e3f2fd;
        border-left: 4px solid #1a73e8;
        padding: 0.6rem 1rem;
        border-radius: 0 8px 8px 0;
        margin: 0.5rem 0;
        font-family: monospace;
        font-size: 0.85rem;
    }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
#  DISTANCE CACHE
# ═══════════════════════════════════════════════════════════════════
class DistanceCache:
    def __init__(self, db_path: str = None):
        self._mem: dict[tuple, float] = {}
        self._write_lock = threading.Lock()
        self.db_path = db_path
        if db_path and Path(db_path).exists():
            self._load_from_db(db_path)

    def _load_from_db(self, db_path):
        try:
            conn = sqlite3.connect(db_path)
            rows = conn.execute("SELECT lat1, lon1, lat2, lon2, dist_km FROM distances").fetchall()
            conn.close()
            for r in rows:
                self._mem[(r[0], r[1], r[2], r[3])] = r[4]
        except Exception:
            pass

    def init_db(self, db_path):
        self.db_path = db_path
        Path(db_path).parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(db_path)
        conn.execute("""CREATE TABLE IF NOT EXISTS distances (
            lat1 REAL, lon1 REAL, lat2 REAL, lon2 REAL, dist_km REAL,
            PRIMARY KEY (lat1, lon1, lat2, lon2))""")
        conn.commit()
        conn.close()

    @staticmethod
    def _key(lat1, lon1, lat2, lon2):
        return (round(float(lat1), 6), round(float(lon1), 6),
                round(float(lat2), 6), round(float(lon2), 6))

    def get(self, lat1, lon1, lat2, lon2):
        return self._mem.get(self._key(lat1, lon1, lat2, lon2))

    def put(self, lat1, lon1, lat2, lon2, dist_km):
        self._mem[self._key(lat1, lon1, lat2, lon2)] = dist_km

    def put_batch(self, items):
        with self._write_lock:
            if self.db_path:
                try:
                    conn = sqlite3.connect(self.db_path)
                    conn.execute("PRAGMA journal_mode=WAL")
                    conn.executemany("INSERT OR IGNORE INTO distances VALUES (?,?,?,?,?)", items)
                    conn.commit()
                    conn.close()
                except Exception:
                    pass
            for lat1, lon1, lat2, lon2, d in items:
                self._mem[self._key(lat1, lon1, lat2, lon2)] = d

    def __len__(self):
        return len(self._mem)


# ═══════════════════════════════════════════════════════════════════
#  DISTANCE FUNCTIONS
# ═══════════════════════════════════════════════════════════════════
def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    lat1, lon1, lat2, lon2 = map(math.radians, map(float, [lat1, lon1, lat2, lon2]))
    a = math.sin((lat2 - lat1) / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin((lon2 - lon1) / 2) ** 2
    return round(R * 2 * math.asin(math.sqrt(a)), 2)


def osrm_table(coords, osrm_url, retries=3):
    coords_str = ";".join(f"{lon},{lat}" for lon, lat in coords)
    url = f"{osrm_url}/table/v1/driving/{coords_str}"
    params = {"annotations": "distance"}
    for attempt in range(retries):
        try:
            r = requests.get(url, params=params, timeout=60)
            if r.status_code == 200:
                data = r.json()
                if data.get("code") == "Ok":
                    return [[round(cell / 1000, 2) if cell is not None else None
                             for cell in row] for row in data["distances"]]
            elif r.status_code == 429:
                time.sleep(2 ** attempt)
        except Exception:
            time.sleep(1)
    return None


def get_distance(cache, lat1, lon1, lat2, lon2):
    cached = cache.get(lat1, lon1, lat2, lon2)
    if cached is not None:
        return cached
    fb = round(haversine(lat1, lon1, lat2, lon2) * 1.3, 2)
    cache.put(lat1, lon1, lat2, lon2, fb)
    return fb


# ═══════════════════════════════════════════════════════════════════
#  PREFETCH DISTANCES
# ═══════════════════════════════════════════════════════════════════
def prefetch_distances(cache, df_points, df_dates, osrm_url, batch_size, api_delay,
                       max_workers, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)

    track_batches = []
    total_pairs = 0
    cached_pairs = 0

    for track in df_points['track'].unique():
        track_pts = df_points[df_points['track'] == track]
        if track_pts.empty:
            continue

        zero_lat = track_pts.iloc[0]['zero_Latitude']
        zero_lon = track_pts.iloc[0]['zero_Longitude']
        if pd.isna(zero_lat) or pd.isna(zero_lon):
            continue

        base = (round(float(zero_lat), 6), round(float(zero_lon), 6))

        point_coords = set()
        for _, row in track_pts.iterrows():
            if pd.isna(row['Latitude']) or pd.isna(row['Longitude']):
                continue
            point_coords.add((round(float(row['Latitude']), 6),
                              round(float(row['Longitude']), 6)))

        all_coords = [base] + [c for c in point_coords if c != base]
        n = len(all_coords)
        if n < 2:
            continue

        n_pairs = n * (n - 1)
        total_pairs += n_pairs
        n_cached = sum(1 for i in range(n) for j in range(n)
                       if i != j and cache.get(all_coords[i][0], all_coords[i][1],
                                                all_coords[j][0], all_coords[j][1]) is not None)
        cached_pairs += n_cached

        if n_cached < n_pairs:
            for start in range(0, n, batch_size):
                chunk = all_coords[start:start + batch_size]
                if len(chunk) >= 2:
                    track_batches.append((track, chunk))

    missing_pairs = total_pairs - cached_pairs
    log(f"Пар потрібно: {total_pairs} | В кеші: {cached_pairs} | Нових: {missing_pairs}")

    if not track_batches:
        log("✅ Всі пари в кеші — API-запити не потрібні!")
        return

    log(f"Table API батчів: {len(track_batches)} (по-треково)")

    def process_batch(batch_coords):
        osrm_coords = [(lon, lat) for lat, lon in batch_coords]
        matrix = osrm_table(osrm_coords, osrm_url)
        items = []
        if matrix is not None:
            for i, (lat1, lon1) in enumerate(batch_coords):
                for j, (lat2, lon2) in enumerate(batch_coords):
                    if i == j:
                        continue
                    dist = matrix[i][j]
                    if dist is not None:
                        items.append((round(lat1, 6), round(lon1, 6),
                                      round(lat2, 6), round(lon2, 6), dist))
        else:
            for i, (lat1, lon1) in enumerate(batch_coords):
                for j, (lat2, lon2) in enumerate(batch_coords):
                    if i == j:
                        continue
                    fb = round(haversine(lat1, lon1, lat2, lon2) * 1.3, 2)
                    items.append((round(lat1, 6), round(lon1, 6),
                                  round(lat2, 6), round(lon2, 6), fb))
        if items:
            cache.put_batch(items)
        time.sleep(api_delay)
        return len(items)

    with ThreadPoolExecutor(max_workers=min(max_workers, 4)) as ex:
        futs = {ex.submit(process_batch, coords): name
                for name, coords in track_batches}
        done = 0
        for f in as_completed(futs):
            done += 1
            try:
                f.result()
                if done % 5 == 0 or done == len(track_batches):
                    log(f"  Table API: {done}/{len(track_batches)} (кеш: {len(cache)})")
            except Exception as e:
                log(f"  ⚠ Помилка: {e}")

    log(f"✅ Prefetch завершено. Кеш: {len(cache)} пар.")


# ═══════════════════════════════════════════════════════════════════
#  OPTIMIZER
# ═══════════════════════════════════════════════════════════════════
def nn_order(base, stop_names, coords, cache):
    rem = list(stop_names)
    ordered = []
    cur = base
    while rem:
        nxt = min(rem, key=lambda s: get_distance(cache, cur[0], cur[1],
                                                    coords[s][0], coords[s][1]))
        ordered.append(nxt)
        cur = coords[nxt]
        rem.remove(nxt)
    return ordered


def calc_km(base, stop_names, coords, cache):
    if not stop_names:
        return 0.0
    ordered = nn_order(base, stop_names, coords, cache)
    pts = [base] + [coords[s] for s in ordered] + [base]
    return round(sum(get_distance(cache, pts[i][0], pts[i][1], pts[i + 1][0], pts[i + 1][1])
                     for i in range(len(pts) - 1)), 2)


def build_library(base, coords, cache, max_stops, n_samples, seed):
    all_stops = list(coords.keys())
    n = len(all_stops)
    rng = random.Random(seed)
    library = {frozenset(): 0.0}

    for stop in all_stops:
        fs = frozenset([stop])
        library[fs] = calc_km(base, [stop], coords, cache)

    actual_max = min(max_stops, n)
    generated = 0
    attempts = 0
    while generated < n_samples and attempts < n_samples * 5:
        attempts += 1
        size = rng.randint(1, actual_max)
        stops = rng.sample(all_stops, size)
        fs = frozenset(stops)
        if fs not in library:
            library[fs] = calc_km(base, stops, coords, cache)
            generated += 1

    return library


def scored_fill(w_days, target_km, library, visited_global, params, seed=None):
    tolerance = params['tolerance']
    max_rest_ratio = params['max_rest_ratio']
    coverage_bonus = params['coverage_bonus']
    repeat_penalty = params['repeat_penalty']
    max_iter = params['max_iter']

    lo, hi = target_km * (1 - tolerance), target_km * (1 + tolerance)
    n_days = len(w_days)
    rng = random.Random(seed)

    lib_sorted = sorted(library.items(), key=lambda x: x[1])
    kms = [v for _, v in lib_sorted]
    fsets = [k for k, _ in lib_sorted]
    n_lib = len(lib_sorted)
    max_rest = int(n_days * max_rest_ratio)

    def try_fill(rng_local, visited_g):
        use_count = {}
        rest_count = 0
        chosen = []
        total = 0.0
        local_visited = set(visited_g)

        for i in range(n_days):
            remaining = n_days - i
            ideal = (target_km - total) / max(remaining, 1)

            window = []
            for j in range(n_lib):
                if kms[j] == 0.0:
                    window.append(j)
                elif ideal * 0.2 <= kms[j] <= ideal * 3.0:
                    window.append(j)
            if len(window) <= 1:
                window = list(range(n_lib))

            weights = []
            for j in window:
                km = kms[j]
                fs = fsets[j]
                km_score = 1.0 / (1.0 + abs(km - ideal) / max(abs(ideal), 0.01))
                has_new = bool(fs) and bool(fs - local_visited)
                cov_bonus = coverage_bonus if has_new else 1.0
                rep = use_count.get(j, 0)
                rep_pen = repeat_penalty ** rep
                rest_pen = 1.0
                if km == 0.0:
                    if rest_count >= max_rest:
                        rest_pen = 0.001
                    elif rest_count >= max_rest * 0.7:
                        rest_pen = 0.1
                weights.append(km_score * cov_bonus * rep_pen * rest_pen)

            tw = sum(weights)
            if tw <= 0:
                tw = 1.0
                weights = [1.0 / len(window)] * len(window)
            r = rng_local.random() * tw
            cum = 0.0
            sel = window[0]
            for j, w in zip(window, weights):
                cum += w
                if r <= cum:
                    sel = j
                    break

            chosen.append(sel)
            total += kms[sel]
            use_count[sel] = use_count.get(sel, 0) + 1
            if kms[sel] == 0.0:
                rest_count += 1
            local_visited |= fsets[sel]

        return chosen, total

    def fine_tune(chosen, total):
        for _ in range(max_iter):
            if lo <= total <= hi:
                break
            mid = (lo + hi) / 2
            if total > hi:
                pos = max(range(n_days), key=lambda i: kms[chosen[i]])
                cur = kms[chosen[pos]]
                if cur == 0.0:
                    break
                need = cur - (total - mid) / max(n_days * 0.1, 1)
                if need <= 0:
                    rest_now = sum(1 for c in chosen if kms[c] == 0.0)
                    if rest_now < max_rest:
                        empty_j = next((j for j in range(n_lib) if kms[j] == 0.0), None)
                        if empty_j is not None:
                            total -= cur
                            chosen[pos] = empty_j
                            continue
                    better = [j for j in range(n_lib) if 0 < kms[j] < cur]
                    if not better:
                        break
                    new_j = min(better, key=lambda j: abs(kms[j] - max(need, 0.01)))
                    total += kms[new_j] - cur
                    chosen[pos] = new_j
                else:
                    better = [j for j in range(n_lib) if 0 < kms[j] < cur]
                    if not better:
                        break
                    new_j = min(better, key=lambda j: abs(kms[j] - need))
                    total += kms[new_j] - cur
                    chosen[pos] = new_j
            else:
                pos = min(range(n_days), key=lambda i: kms[chosen[i]])
                cur = kms[chosen[pos]]
                need = cur + (mid - total) / max(n_days * 0.1, 1)
                better = [j for j in range(n_lib) if kms[j] > cur]
                if not better:
                    break
                new_j = min(better, key=lambda j: abs(kms[j] - need))
                total += kms[new_j] - cur
                chosen[pos] = new_j
        return chosen, total

    best_chosen = None
    best_var = -1
    best_total = None
    for _ in range(40):
        tr = random.Random(rng.randint(0, 10 ** 9))
        ch, tot = try_fill(tr, visited_global)
        ch, tot = fine_tune(ch[:], tot)
        var = len(set(ch))
        if lo <= tot <= hi:
            if var > best_var:
                best_var = var
                best_chosen = ch[:]
                best_total = tot
        elif best_chosen is None:
            best_var = var
            best_chosen = ch[:]
            best_total = tot

    day_plan = {}
    new_visited = set(visited_global)
    for day, idx in zip(w_days, best_chosen):
        day_plan[day] = (fsets[idx], kms[idx])
        new_visited |= fsets[idx]

    return day_plan, new_visited


def build_rows(track, date, stop_names, coords, cache, owner_bases, date_owner_map,
               default_base, addresses=None):
    """Build output rows for one day.

    owner_bases    — {OwnerName: (lat, lon, zero_adress_str)}
    date_owner_map — {date -> OwnerName} for this track
    default_base   — fallback (lat, lon) if owner not in owner_bases
    addresses      — {FullName -> address_str} appended after | in Точка column
    """
    COL_TRACK = "Трек"
    COL_DATE  = "Дата"
    COL_OWNER = "OwnerName"
    COL_POINT = "Точка"
    COL_KM    = "Пробег_км"
    COL_LAT   = "Latitude"
    COL_LON   = "Longitude"

    # Determine OwnerName for this date
    date_key = pd.Timestamp(date).date()
    owner = date_owner_map.get(date_key, "")

    # Resolve base coordinates and home label for this specific owner/day
    if owner and owner in owner_bases:
        b_lat, b_lon, zero_adress = owner_bases[owner]
        base = (b_lat, b_lon)
    else:
        base = default_base
        zero_adress = ""
    home_label = zero_adress if zero_adress else "База"

    if not stop_names:
        return [{
            COL_TRACK: track, COL_DATE: date.strftime("%d.%m.%Y"),
            COL_OWNER: owner,
            COL_POINT: "— без виїзду —", COL_KM: 0.0,
            COL_LAT: base[0], COL_LON: base[1]
        }]

    ordered = nn_order(base, list(stop_names), coords, cache)
    pts = [base] + [coords[s] for s in ordered] + [base]

    def point_label(name):
        if addresses:
            addr = addresses.get(name, "")
            return f"{name} | {addr}" if addr else name
        return name

    names = [f"СТАРТ: {home_label}"] + [point_label(s) for s in ordered] + [f"ФІНІШ: {home_label}"]
    rows = []
    for i, (name, (lat, lon)) in enumerate(zip(names, pts)):
        km = 0.0 if i == 0 else round(
            get_distance(cache, pts[i - 1][0], pts[i - 1][1], lat, lon), 2)
        if km == 0.0 and i > 0 and not name.startswith("СТАРТ"):
            continue
        rows.append({
            COL_TRACK: track, COL_DATE: date.strftime("%d.%m.%Y"),
            COL_OWNER: owner,
            COL_POINT: name, COL_KM: km,
            COL_LAT: lat, COL_LON: lon
        })
    return rows


def extract_pool(df_points_track):
    """Extract stop coords, per-owner bases and address map from a track's points.

    Returns:
        coords       — {FullName: (lat, lon)}
        default_base — centroid of all owner bases (used for build_library)
        owner_bases  — {OwnerName: (lat, lon, zero_adress_str)}
        addresses    — {FullName: address_str}
    """
    coords = {}
    addresses = {}
    owner_bases = {}  # {OwnerName: (lat, lon, adress_str)}

    for _, row in df_points_track.iterrows():
        lat, lon = row['Latitude'], row['Longitude']
        if pd.isna(lat) or pd.isna(lon):
            continue
        name = str(row['FullName']).strip()
        coords[name] = (float(lat), float(lon))
        addr = row.get('adress FullName', row.get('address', ""))
        addresses[name] = (str(addr).strip()
                           if addr is not None and not (isinstance(addr, float) and math.isnan(addr))
                           else "")

    # Per-owner home base
    if 'OwnerName' in df_points_track.columns:
        for owner, grp in df_points_track.groupby('OwnerName'):
            row0 = grp.iloc[0]
            z_lat = row0.get('zero_Latitude')
            z_lon = row0.get('zero_Longitude')
            if z_lat is not None and z_lon is not None and not pd.isna(z_lat) and not pd.isna(z_lon):
                za = row0.get('zero_adress', row0.get('zero_address', ""))
                za_str = (str(za).strip()
                          if za is not None and not (isinstance(za, float) and math.isnan(za))
                          else "")
                owner_bases[str(owner)] = (float(z_lat), float(z_lon), za_str)

    # Default base — centroid of all owner bases (fallback: centroid of points)
    if owner_bases:
        lats = [v[0] for v in owner_bases.values()]
        lons = [v[1] for v in owner_bases.values()]
        default_base = (sum(lats) / len(lats), sum(lons) / len(lons))
    elif coords:
        lats = [v[0] for v in coords.values()]
        lons = [v[1] for v in coords.values()]
        default_base = (sum(lats) / len(lats), sum(lons) / len(lons))
    else:
        default_base = None

    return coords, default_base, owner_bases, addresses


def parse_period(period):
    parts = str(period).strip().replace("/", "-").replace(".", "-").split("-")
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) != 2:
        raise ValueError(f"Невідомий формат: {period}")
    return (int(parts[0]), int(parts[1])) if len(parts[0]) == 4 else (int(parts[1]), int(parts[0]))


def working_days(year, month):
    return list(pd.bdate_range(
        start=f"{year}-{month:02d}-01",
        end=pd.Timestamp(year, month, 1) + pd.offsets.MonthEnd(0),
        freq="B"))


def compute_odometers(summary_rows_for_track, odometr_last):
    """
    Compute odometer values for each period in a track.

    summary_rows_for_track: list of dicts with keys 'period', 'fact' — sorted chronologically
    odometr_last: float odometer value set for the last period (END of last period)

    Returns: dict {period_str -> odometer_value}
    
    Logic:
    - The odometr value in sh3 is the reading at the END of the last period.
    - For earlier periods:  odo_end[i] = odo_end[i+1] - fact_km[i+1]
    - For single period: just show the odometr value as-is.
    """
    if odometr_last is None or (isinstance(odometr_last, float) and math.isnan(odometr_last)):
        return {}

    if len(summary_rows_for_track) == 1:
        return {summary_rows_for_track[0]['period']: round(odometr_last, 1)}

    # Sort chronologically
    def period_key(s):
        try:
            y, m = parse_period(s['period'])
            return (y, m)
        except Exception:
            return (0, 0)

    sorted_rows = sorted(summary_rows_for_track, key=period_key)
    odo_map = {}
    # Last period gets the provided odometr value
    odo_map[sorted_rows[-1]['period']] = round(odometr_last, 1)

    # Walk backwards
    for i in range(len(sorted_rows) - 2, -1, -1):
        next_period = sorted_rows[i + 1]['period']
        next_fact = sorted_rows[i + 1]['fact']
        prev_odo = odo_map[next_period]
        odo_map[sorted_rows[i]['period']] = round(prev_odo - next_fact, 1)

    return odo_map


def write_xlsx_to_bytes(all_rows, summary_rows, tolerance, odometr_map):
    """
    odometr_map: dict { (track, period) -> odometer_value }
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Маршрути"
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    hfl = PatternFill("solid", start_color="2F5496")
    thin = Side(style="thin", color="BBBBBB")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    sf = PatternFill("solid", start_color="E2EFDA")
    ef = PatternFill("solid", start_color="FCE4D6")
    rf = PatternFill("solid", start_color="D9E2F3")
    own_f = PatternFill("solid", start_color="FFF9C4")  # light yellow for owner col

    # Маршрути headers — added OwnerName
    headers = ["Трек", "Дата", "OwnerName", "Точка", "Пробег_км", "Latitude", "Longitude"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(1, c, h)
        cell.font = hf
        cell.fill = hfl
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(all_rows, 2):
        for c, key in enumerate(headers, 1):
            val = row.get(key, "")
            cell = ws1.cell(r, c, val)
            cell.font = Font(name="Arial", size=9)
            cell.border = brd
            cell.alignment = Alignment(horizontal="left" if c < 5 else "center")
            pt = row.get("Точка", "")
            if pt.startswith("СТАРТ:"):
                cell.fill = sf
            elif pt.startswith("ФІНІШ:"):
                cell.fill = ef
            elif pt == "— без виїзду —":
                cell.fill = rf
            elif c == 3:  # OwnerName column — subtle highlight
                cell.fill = own_f

    for i, w in enumerate([38, 12, 22, 65, 11, 13, 13], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Зведення sheet — added Одометр
    ws2 = wb.create_sheet("Зведення")
    sh = ["Трек", "Період", "План, км", "Факт, км", "Відхилення, %", "Статус", "Одометр"]
    for c, h in enumerate(sh, 1):
        cell = ws2.cell(1, c, h)
        cell.font = hf
        cell.fill = hfl
        cell.alignment = Alignment(horizontal="center")

    ok_f = PatternFill("solid", start_color="C6EFCE")
    er_f = PatternFill("solid", start_color="FFC7CE")
    odo_f = PatternFill("solid", start_color="E8EAF6")  # indigo-50 for odometer col
    ok_n = Font(name="Arial", size=9, color="006100")
    er_n = Font(name="Arial", size=9, color="9C0006")

    for r, s in enumerate(summary_rows, 2):
        ok = abs(s["dev_pct"]) <= tolerance * 100
        odo_val = odometr_map.get((s["track"], s["period"]), "")
        odo_str = f"{odo_val:,.1f}" if isinstance(odo_val, (int, float)) else ""

        vals = [
            s["track"], s["period"], s["plan"], round(s["fact"], 1),
            f'{s["dev_pct"]:+.1f}%',
            "✓ В межах" if ok else "✗ Поза межами",
            odo_str
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(r, c, v)
            cell.border = brd
            cell.alignment = Alignment(horizontal="center")
            if c in (5, 6):
                cell.font = ok_n if ok else er_n
                cell.fill = ok_f if ok else er_f
            elif c == 7:
                cell.font = Font(name="Arial", size=9, bold=True, color="283593")
                cell.fill = odo_f
            else:
                cell.font = Font(name="Arial", size=9)

    for i, w in enumerate([38, 14, 12, 12, 18, 20, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════
#  MAIN OPTIMIZATION PIPELINE
# ═══════════════════════════════════════════════════════════════════
def run_optimization(df_points, df_dates, df_plan, params, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)

    # Initialize cache
    cache = DistanceCache()
    tmp_db = os.path.join(tempfile.gettempdir(), "route_opt_cache.db")
    cache.init_db(tmp_db)
    log(f"Кеш: {tmp_db}")

    # Step 1: Prefetch distances
    log("\n═══ КРОК 1: Завантаження відстаней (OSRM) ═══")
    prefetch_distances(
        cache, df_points, df_dates,
        osrm_url=params['osrm_url'],
        batch_size=params['batch_size'],
        api_delay=params['api_delay'],
        max_workers=params['max_workers'],
        log_callback=log
    )

    # Step 2: Optimize routes
    log("\n═══ КРОК 2: Оптимізація маршрутів ═══")

    # Build date→owner lookup per track from sh2
    # {track: {date: OwnerName}}
    date_owner_by_track = {}
    if 'OwnerName' in df_dates.columns:
        for _, row in df_dates.iterrows():
            track = row['track']
            try:
                d = pd.to_datetime(row['DateActivity']).date()
            except Exception:
                continue
            owner = str(row.get('OwnerName', "")).strip()
            date_owner_by_track.setdefault(track, {})[d] = owner
    
    # Build odometr lookup from sh3 — support both 'одометр' and 'odometr' column names
    odo_col = None
    for candidate in ['одометр', 'odometr']:
        if candidate in df_plan.columns:
            odo_col = candidate
            break
    odometr_input = {}
    if odo_col:
        for _, row in df_plan.iterrows():
            track = str(row['Трек']).strip()
            odo = row.get(odo_col, None)
            if odo is not None and not (isinstance(odo, float) and math.isnan(odo)):
                odometr_input.setdefault(track, []).append((str(row['Період']).strip(), float(odo)))

    # Filter active plans
    df_plan_active = df_plan[df_plan['План, км'] > 0].drop_duplicates(
        subset=['Трек', 'Період']).copy()
    tracks = df_plan_active['Трек'].unique()
    log(f"Активних пар Трек+Період: {len(df_plan_active)}")

    all_rows_out = []
    summary_rows = []

    for track in tracks:
        log(f"\n{'═' * 50}\nТРЕК: {track}")

        track_pts = df_points[df_points['track'] == track].copy()
        if track_pts.empty:
            log("  ⚠ Немає точок — пропускаємо.")
            continue

        # Date→owner map for this track
        date_owner_map = date_owner_by_track.get(track, {})

        # ── Per-owner point pools ────────────────────────────────
        # Each owner has their own territory (points) and home base.
        # We build a separate library per owner so routes stay within
        # each driver's geographical area.
        owner_pool = {}   # {owner: (coords, base, addresses)}
        for own in set(date_owner_map.values()) - {""}:
            own_pts = track_pts[track_pts['OwnerName'] == own]
            if own_pts.empty:
                # owner drives the track but has no points in sh1 for it — skip
                continue
            c_o, b_o, ob_o, a_o = extract_pool(own_pts)
            if c_o and b_o is not None:
                owner_pool[own] = (c_o, b_o, {own: ob_o.get(own, ob_o.get(list(ob_o.keys())[0])) if ob_o else (b_o[0], b_o[1], "")}, a_o)
                log(f"  Водій: {own} | точок: {len(c_o)} | база: {b_o[0]:.4f},{b_o[1]:.4f}")

        if not owner_pool:
            # Fallback: single shared pool (original behaviour)
            coords, default_base, owner_bases, addresses = extract_pool(track_pts)
            if not coords or default_base is None:
                log("  ⚠ Немає координат — пропускаємо.")
                continue
            owner_pool = {'__all__': (coords, default_base, owner_bases, addresses)}
            log(f"  Пул точок (спільний): {len(coords)}")

        # Get periods for this track
        track_plans = df_plan_active[df_plan_active['Трек'] == track]
        months_info = []
        for _, row in track_plans.iterrows():
            period = str(row['Період']).strip()
            target = float(str(row['План, км']).replace(",", "."))
            try:
                year, month = parse_period(period)
            except ValueError as e:
                log(f"  ⚠ {e}")
                continue

            track_dates = df_dates[df_dates['track'] == track].copy()
            track_dates['DateActivity'] = pd.to_datetime(track_dates['DateActivity'])
            track_dates['Date'] = track_dates['DateActivity'].dt.date

            month_dates = track_dates[
                (track_dates['DateActivity'].dt.year == year) &
                (track_dates['DateActivity'].dt.month == month)
            ]['Date'].unique()

            w_days = sorted([pd.Timestamp(d) for d in month_dates]) if len(month_dates) > 0 else working_days(year, month)
            months_info.append((year, month, target, w_days, period))

        if not months_info:
            continue

        months_info.sort(key=lambda x: (x[0], x[1]))
        log(f"  Місяців: {len(months_info)}, днів: {sum(len(m[3]) for m in months_info)}")

        # Build one library per owner
        owner_libraries = {}
        for own, (c_o, b_o, ob_o, a_o) in owner_pool.items():
            lib = build_library(b_o, c_o, cache,
                                params['max_stops'], params['n_samples'], params['seed'])
            owner_libraries[own] = lib
            log(f"  Бібліотека [{own}]: {len(lib)} маршрутів")

        # Combine all owner bases into one dict for build_rows
        all_owner_bases = {}
        all_addresses = {}
        for own, (c_o, b_o, ob_o, a_o) in owner_pool.items():
            all_owner_bases.update(ob_o)
            all_addresses.update(a_o)
        all_coords = {k: v for c_o, _, _, _ in owner_pool.values() for k, v in c_o.items()}

        visited_per_owner = {own: set() for own in owner_pool}
        track_summary_rows = []

        for year, month, target, w_days, period in months_info:
            log(f"\n  ── Період: {period} | План: {target} км | Днів: {len(w_days)}")

            # Split days by owner
            owner_days = {}
            for d in w_days:
                own = date_owner_map.get(d.date(), "")
                # Match to an owner pool key
                pool_key = own if own in owner_pool else ('__all__' if '__all__' in owner_pool else None)
                if pool_key:
                    owner_days.setdefault(pool_key, []).append(d)

            if not owner_days:
                log("  ⚠ Немає днів — пропускаємо.")
                continue

            # Allocate target km proportionally to number of days per owner
            total_days = sum(len(v) for v in owner_days.values())
            combined_day_plan = {}
            fact_km = 0.0

            for own, o_days in owner_days.items():
                c_o, b_o, ob_o, a_o = owner_pool[own]
                lib_o = owner_libraries[own]
                o_target = round(target * len(o_days) / total_days, 2)
                log(f"    {own}: {len(o_days)} днів | ціль {o_target} км")

                o_plan, visited_per_owner[own] = scored_fill(
                    o_days, o_target, lib_o, visited_per_owner[own], params
                )
                combined_day_plan.update(o_plan)

                for day_ts, (fs, _) in o_plan.items():
                    fact_km += calc_km(b_o, list(fs), c_o, cache)

            fact_km = round(fact_km, 2)
            dev_pct = (fact_km - target) / target * 100
            status = "✓" if abs(dev_pct) <= params['tolerance'] * 100 else "✗"
            log(f"  Факт: {fact_km:.1f} км | Відхилення: {dev_pct:+.1f}% {status}")

            covered_now = set()
            for fs, _ in combined_day_plan.values():
                covered_now |= fs
            log(f"  Унікальних точок: {len(covered_now)}/{len(all_coords)}")

            rest_days = sum(1 for _, (fs, _) in combined_day_plan.items() if not fs)
            log(f"  Днів з виїздом: {len(combined_day_plan) - rest_days}, без виїзду: {rest_days}")

            for date in sorted(combined_day_plan.keys()):
                fs, _ = combined_day_plan[date]
                # Use per-owner coords for this day
                day_owner = date_owner_map.get(date.date(), "")
                pool_key = day_owner if day_owner in owner_pool else '__all__'
                c_day = owner_pool[pool_key][0] if pool_key in owner_pool else all_coords
                all_rows_out.extend(
                    build_rows(track, pd.Timestamp(date), fs, c_day, cache,
                               all_owner_bases, date_owner_map,
                               owner_pool[pool_key][1] if pool_key in owner_pool else list(all_owner_bases.values())[0][:2],
                               all_addresses))

            row_entry = {
                "track": track, "period": period,
                "plan": target, "fact": fact_km, "dev_pct": dev_pct
            }
            summary_rows.append(row_entry)
            track_summary_rows.append(row_entry)

        total_covered = sum(len(v) for v in visited_per_owner.values())
        total_pts = len(all_coords)
        log(f"\n  Покрито за всі місяці: {total_covered}/{total_pts}")

    if not all_rows_out:
        log("\n⚠ Немає даних для запису!")
        return None, None, summary_rows, {}

    # ── Compute odometer map ──
    # Group summary rows by track
    odometr_map = {}  # {(track, period) -> odo_value}

    tracks_in_summary = {}
    for s in summary_rows:
        tracks_in_summary.setdefault(s['track'], []).append(s)

    for track, rows in tracks_in_summary.items():
        # Find odometr value for last period from input
        odo_entries = odometr_input.get(track, [])
        if not odo_entries:
            continue

        # Sort periods in rows chronologically
        def period_key_fn(s):
            try:
                y, m = parse_period(s['period'])
                return (y, m)
            except Exception:
                return (0, 0)

        sorted_rows = sorted(rows, key=period_key_fn)

        # Determine which period has the odometr value
        # The odometr in sh3 is set for the "last" period of the track
        # We pick the entry that matches the last period (or just use the one provided)
        last_period_str = sorted_rows[-1]['period']

        # Try to find a matching odometr entry for last period
        odo_val = None
        for p, v in odo_entries:
            if p == last_period_str:
                odo_val = v
                break
        # If no explicit match, use any provided value (assume it's for last period)
        if odo_val is None and odo_entries:
            odo_val = odo_entries[0][1]

        odo_per_period = compute_odometers(sorted_rows, odo_val)
        for period, odo in odo_per_period.items():
            odometr_map[(track, period)] = odo

    # Build output
    xlsx_bytes = write_xlsx_to_bytes(all_rows_out, summary_rows, params['tolerance'], odometr_map)

    ok_cnt = sum(1 for s in summary_rows if abs(s["dev_pct"]) <= params['tolerance'] * 100)
    log(f"\n{'═' * 50}")
    log(f"РЕЗУЛЬТАТ: {len(summary_rows)} періодів | ✓ {ok_cnt} | ✗ {len(summary_rows) - ok_cnt}")

    return xlsx_bytes, all_rows_out, summary_rows, odometr_map


# ═══════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════

def render_header():
    st.markdown('<div class="main-title">🚗 Route Optimizer v7</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="sub-title">Оптимізація маршрутів за планом пробігу • '
        'OSRM + Monte Carlo • Мульти-власник • Одометр</div>',
        unsafe_allow_html=True)


def render_sidebar():
    with st.sidebar:
        st.markdown("### Параметри")
        st.markdown("---")

        st.markdown("**Оптимізація**")
        tolerance = st.slider("TOLERANCE — допуск від плану",
                              0.01, 0.30, 0.10, 0.01)
        max_stops = st.slider("MAX_STOPS — макс. точок/день", 1, 15, 5)
        n_samples = st.slider("N_SAMPLES — розмір бібліотеки", 500, 15000, 3000, 500)
        seed = st.number_input("SEED — зерно випадковості", value=42, min_value=0, step=1)
        max_iter = st.slider("MAX_ITER — ітерації fine-tune", 1000, 20000, 8000, 1000)

        st.markdown("---")
        st.markdown("**Балансування**")
        coverage_bonus = st.slider("COVERAGE_BONUS — бонус за нові точки", 1.0, 15.0, 3.0, 0.5)
        repeat_penalty = st.slider("REPEAT_PENALTY — штраф за повтори", 0.0, 1.0, 0.3, 0.05)
        max_rest_ratio = st.slider("MAX_REST_RATIO — ліміт днів без виїзду", 0.0, 0.9, 0.6, 0.05)

        st.markdown("---")
        st.markdown("**OSRM API**")
        osrm_url = st.text_input("OSRM URL", value="http://router.project-osrm.org")
        batch_size = st.slider("Batch size (Table API)", 10, 100, 80)
        api_delay = st.slider("API delay (сек)", 0.05, 1.0, 0.15, 0.05)
        max_workers = st.slider("Потоки", 1, 16, 8)

    return {
        'tolerance': tolerance,
        'max_stops': max_stops,
        'n_samples': n_samples,
        'seed': seed,
        'max_iter': max_iter,
        'coverage_bonus': coverage_bonus,
        'repeat_penalty': repeat_penalty,
        'max_rest_ratio': max_rest_ratio,
        'osrm_url': osrm_url.rstrip('/'),
        'batch_size': batch_size,
        'api_delay': api_delay,
        'max_workers': max_workers,
    }


def render_map_editor(df_points):
    """Render interactive map for viewing and editing coordinates."""
    try:
        import folium
        from streamlit_folium import st_folium
    except ImportError:
        st.warning("Встановіть folium та streamlit-folium для роботи з картою.")
        return df_points

    st.markdown("#### Координати точок")

    missing_point = df_points[df_points['Latitude'].isna() | df_points['Longitude'].isna()]
    missing_base = df_points[df_points['zero_Latitude'].isna() | df_points['zero_Longitude'].isna()]
    has_missing = len(missing_point) > 0 or len(missing_base) > 0

    if has_missing:
        st.warning(f"Відсутні координати: {len(missing_point)} точок, {len(missing_base)} баз.")

    col_mode, col_track, col_target = st.columns([1, 1.5, 2.5])

    with col_mode:
        edit_mode = st.selectbox(
            "Що редагувати",
            ["Координати точки", "Координати бази"],
            key="map_edit_mode"
        )

    tracks_list = df_points['track'].unique().tolist()
    with col_track:
        selected_track = st.selectbox("Трек", tracks_list, key="map_track")

    with col_target:
        if edit_mode == "Координати точки":
            track_pts = df_points[df_points['track'] == selected_track]
            point_options = []
            for _, row in track_pts.iterrows():
                name = str(row['FullName'])
                has_coords = not pd.isna(row['Latitude']) and not pd.isna(row['Longitude'])
                label = name if has_coords else f"[немає координат] {name}"
                point_options.append(label)
            selected_point_label = st.selectbox("Точка", point_options, key="map_point")
            selected_point = selected_point_label.replace("[немає координат] ", "")
        else:
            zero_lat = df_points[df_points['track'] == selected_track].iloc[0].get('zero_Latitude')
            zero_lon = df_points[df_points['track'] == selected_track].iloc[0].get('zero_Longitude')
            if pd.isna(zero_lat) or pd.isna(zero_lon):
                st.info("База не визначена — клікніть на карті")
            else:
                st.info(f"Поточна база: {zero_lat:.6f}, {zero_lon:.6f}")

    track_pts = df_points[df_points['track'] == selected_track]
    valid_pts = track_pts.dropna(subset=['Latitude', 'Longitude'])

    z_lat = track_pts.iloc[0].get('zero_Latitude') if len(track_pts) > 0 else None
    z_lon = track_pts.iloc[0].get('zero_Longitude') if len(track_pts) > 0 else None

    if not pd.isna(z_lat) and not pd.isna(z_lon):
        center_lat, center_lon = float(z_lat), float(z_lon)
        zoom = 10
    elif len(valid_pts) > 0:
        center_lat = valid_pts['Latitude'].mean()
        center_lon = valid_pts['Longitude'].mean()
        zoom = 10
    else:
        center_lat, center_lon = 49.0, 32.0
        zoom = 6

    m = folium.Map(location=[center_lat, center_lon], zoom_start=zoom,
                   tiles="CartoDB positron")

    if not pd.isna(z_lat) and not pd.isna(z_lon):
        folium.Marker(
            [z_lat, z_lon],
            popup=f"<b>База:</b> {selected_track[:30]}",
            tooltip=f"База: {selected_track[:25]}",
            icon=folium.Icon(color="green", icon="home", prefix="fa"),
        ).add_to(m)

    for _, row in track_pts.iterrows():
        if pd.isna(row['Latitude']) or pd.isna(row['Longitude']):
            continue
        name = str(row['FullName'])
        folium.CircleMarker(
            [row['Latitude'], row['Longitude']],
            radius=7, color="#1a73e8", fill=True,
            fill_color="#1a73e8", fill_opacity=0.85,
            popup=f"<b>{name[:50]}</b><br>{row['Latitude']:.6f}, {row['Longitude']:.6f}",
            tooltip=name[:35],
        ).add_to(m)

    map_data = st_folium(m, width=None, height=450, returned_objects=["last_clicked"])

    if map_data and map_data.get("last_clicked"):
        clicked = map_data["last_clicked"]
        lat_c = round(clicked['lat'], 6)
        lon_c = round(clicked['lng'], 6)

        st.markdown(
            f'<div class="coord-info">Клік на карті: <b>{lat_c}</b>, <b>{lon_c}</b></div>',
            unsafe_allow_html=True)

        col_btn, col_desc = st.columns([1, 3])
        with col_btn:
            apply_clicked = st.button("Застосувати", type="primary",
                                       use_container_width=True, key="btn_apply_coord")
        with col_desc:
            if edit_mode == "Координати точки":
                st.caption(f"→ {selected_point[:60]} ({selected_track[:25]})")
            else:
                st.caption(f"→ База для {selected_track[:40]}")

        if apply_clicked:
            if edit_mode == "Координати точки":
                idx = df_points[
                    (df_points['track'] == selected_track) &
                    (df_points['FullName'] == selected_point)
                ].index
                if len(idx) > 0:
                    df_points.loc[idx, 'Latitude'] = lat_c
                    df_points.loc[idx, 'Longitude'] = lon_c
                    st.success(f"Координати оновлено: {selected_point[:50]}")
                    st.rerun()
                else:
                    st.error("Точку не знайдено")
            else:
                idx = df_points[df_points['track'] == selected_track].index
                if len(idx) > 0:
                    df_points.loc[idx, 'zero_Latitude'] = lat_c
                    df_points.loc[idx, 'zero_Longitude'] = lon_c
                    st.success(f"Базу оновлено для: {selected_track[:40]}")
                    st.rerun()
                else:
                    st.error("Трек не знайдено")
    else:
        st.caption("Клікніть на карті щоб обрати координати")

    return df_points


def main():
    render_header()
    params = render_sidebar()

    uploaded = st.file_uploader(
        "📂 Завантажте вхідний файл (.xlsx)",
        type=["xlsx"],
        help=(
            "Файл з листами: sh1 (точки), sh2 (дати), sh3 (план пробігу)\n\n"
            "**sh1** колонки: track, OwnerName, FullName, address, Latitude, Longitude, "
            "zero_Latitude, zero_Longitude, zero_address\n\n"
            "**sh2** колонки: DateActivity, track, OwnerName\n\n"
            "**sh3** колонки: Трек, Період, План км, odometr (тільки для останнього Період кожного Трек)"
        )
    )

    if uploaded is None:
        st.info(
            "Завантажте файл для початку роботи.\n\n"
            "**Нові колонки у v7:**\n"
            "- **sh1**: `address` (адреса точки доставки), `zero_address` (адреса дому водія)\n"
            "- **sh2**: `OwnerName` (хто їде в цей день; одночасно лише один водій на трек)\n"
            "- **sh3**: `odometr` (показання одометра для останнього Період кожного Трек)\n\n"
            "**Вихід:**\n"
            "- Маршрути: СТАРТ/ФІНІШ показують адресу водія з `zero_address`\n"
            "- Зведення: стовпець **Одометр** розраховується назад від останнього Період"
        )
        return

    try:
        xls = pd.ExcelFile(uploaded)
        missing_sheets = [s for s in ['sh1', 'sh2', 'sh3'] if s not in xls.sheet_names]
        if missing_sheets:
            st.error(f"Файл має містити листи: sh1, sh2, sh3. Відсутні: {missing_sheets}")
            return

        df_points = pd.read_excel(xls, sheet_name="sh1")
        df_dates = pd.read_excel(xls, sheet_name="sh2")
        df_plan = pd.read_excel(xls, sheet_name="sh3")
    except Exception as e:
        st.error(f"Помилка читання файлу: {e}")
        return

    # ── Normalize whitespace in all string columns that carry track names ──
    # Excel sometimes stores non-breaking spaces (\xa0) instead of regular spaces,
    # causing track name mismatches between sh1/sh2/sh3.
    def normalize_ws(df, cols):
        for col in cols:
            if col in df.columns:
                df[col] = (df[col].astype(str)
                           .str.replace('\xa0', ' ', regex=False)
                           .str.replace('\u00a0', ' ', regex=False)
                           .str.strip())
        return df

    df_points  = normalize_ws(df_points,  ['OwnerName', 'track'])
    df_dates   = normalize_ws(df_dates,   ['track', 'OwnerName'])
    df_plan    = normalize_ws(df_plan,    ['Трек'])

    # Ensure new columns exist (backward compat)
    for col in ['adress FullName', 'zero_adress']:
        if col not in df_points.columns:
            df_points[col] = ""
    if 'OwnerName' not in df_dates.columns:
        df_dates['OwnerName'] = ""
    odo_col_exists = 'одометр' in df_plan.columns or 'odometr' in df_plan.columns
    if not odo_col_exists:
        df_plan['одометр'] = None

    # ── Derive `track` for sh1 from sh2 (OwnerName → track mapping) ──
    # sh1 has no `track` column; each point belongs to all tracks the owner drives.
    if 'track' not in df_points.columns:
        owner_track_pairs = (df_dates[['OwnerName', 'track']]
                             .dropna()
                             .drop_duplicates())
        # Expand: one row per (point × track) for owners that drive multiple tracks
        df_points = df_points.merge(owner_track_pairs, on='OwnerName', how='left')
        missing_track = df_points['track'].isna().sum()
        if missing_track > 0:
            st.warning(f"⚠️ {missing_track} точок не мають відповідного треку в sh2 — перевірте OwnerName.")
            df_points = df_points.dropna(subset=['track'])

    if 'df_points' not in st.session_state or st.session_state.get('_file_id') != uploaded.name:
        st.session_state.df_points = df_points.copy()
        st.session_state._file_id = uploaded.name

    df_points = st.session_state.df_points

    # ── Stats row ──
    tracks = df_points['track'].unique()
    cols = st.columns(4)
    with cols[0]:
        st.markdown(f'<div class="stat-card"><h3>{len(tracks)}</h3><p>Треків</p></div>',
                    unsafe_allow_html=True)
    with cols[1]:
        st.markdown(f'<div class="stat-card"><h3>{len(df_points)}</h3><p>Точок</p></div>',
                    unsafe_allow_html=True)
    with cols[2]:
        n_dates = len(pd.to_datetime(df_dates['DateActivity']).dt.date.unique())
        st.markdown(f'<div class="stat-card"><h3>{n_dates}</h3><p>Унікальних дат</p></div>',
                    unsafe_allow_html=True)
    with cols[3]:
        # Count unique owners across sh1
        all_owners = df_points['OwnerName'].dropna().unique() if 'OwnerName' in df_points.columns else []
        st.markdown(f'<div class="stat-card"><h3>{len(all_owners)}</h3><p>Водіїв</p></div>',
                    unsafe_allow_html=True)

    st.markdown("")

    tab_map, tab_pts, tab_dates, tab_plan = st.tabs([
        "Карта та координати", "Точки (sh1)", "Дати (sh2)", "План (sh3)"
    ])

    with tab_map:
        df_points = render_map_editor(df_points)
        st.session_state.df_points = df_points

    with tab_pts:
        st.markdown("#### Точки відвідування")
        # Show columns including new address fields
        edit_cols = ['track', 'OwnerName', 'FullName', 'adress FullName',
                     'Latitude', 'Longitude',
                     'zero_Latitude', 'zero_Longitude', 'zero_adress']
        display_cols = [c for c in edit_cols if c in df_points.columns]
        edited_pts = st.data_editor(
            df_points[display_cols],
            use_container_width=True,
            num_rows="dynamic",
            key="pts_editor"
        )
        if st.button("💾 Зберегти зміни координат", key="save_pts"):
            for c in display_cols:
                if c in edited_pts.columns:
                    st.session_state.df_points.loc[edited_pts.index, c] = edited_pts[c]
            st.success("Дані оновлено!")

        # Show owner-track mapping info
        if 'OwnerName' in df_points.columns:
            owner_track = df_points.groupby('track')['OwnerName'].apply(
                lambda x: ', '.join(sorted(x.dropna().unique()))).reset_index()
            owner_track.columns = ['Трек', 'Водії']
            st.markdown("**Розподіл водіїв по треках:**")
            st.dataframe(owner_track, use_container_width=True, hide_index=True)

    with tab_dates:
        st.markdown("#### Дати виїздів (sh2)")
        df_dates_show = df_dates.copy()
        df_dates_show['DateActivity'] = pd.to_datetime(df_dates_show['DateActivity'])

        # Summary grouped by track + owner
        group_cols = ['track']
        if 'OwnerName' in df_dates_show.columns:
            group_cols.append('OwnerName')
        date_summary = df_dates_show.groupby(group_cols).agg(
            Днів=('DateActivity', 'nunique'),
            Перша_дата=('DateActivity', 'min'),
            Остання_дата=('DateActivity', 'max'),
        ).reset_index()
        date_summary['Перша_дата'] = date_summary['Перша_дата'].dt.strftime('%Y-%m-%d')
        date_summary['Остання_дата'] = date_summary['Остання_дата'].dt.strftime('%Y-%m-%d')
        st.dataframe(date_summary, use_container_width=True)

        # Conflict check: same track, same day, multiple owners
        if 'OwnerName' in df_dates.columns:
            df_dates_tmp = df_dates.copy()
            df_dates_tmp['_date'] = pd.to_datetime(df_dates_tmp['DateActivity']).dt.date
            conflict = df_dates_tmp.groupby(['track', '_date'])['OwnerName'].nunique()
            conflicts = conflict[conflict > 1]
            if len(conflicts) > 0:
                st.error(
                    f"⚠️ Знайдено {len(conflicts)} конфлікт(ів): один трек — кілька водіїв в той самий день!\n"
                    f"Перевірте sh2 — одночасно лише один OwnerName може використовувати трек."
                )
                st.dataframe(conflicts.reset_index().rename(
                    columns={'_date': 'Дата', 'OwnerName': 'Кількість водіїв'}),
                    use_container_width=True)

    with tab_plan:
        st.markdown("#### Плановий пробіг (sh3)")
        st.markdown(
            "Стовпець **`odometr`** заповнюється **лише для останнього Період** кожного Трека. "
            "Показання для попередніх Período розраховуються автоматично: "
            "`Одометр[i] = Одометр[i+1] − Факт_км[i+1]`"
        )
        plan_edit_cols = ['Трек', 'Період', 'План, км', 'odometr']
        plan_display = [c for c in plan_edit_cols if c in df_plan.columns]
        st.dataframe(df_plan[plan_display], use_container_width=True)

    # ── Run button ──
    st.markdown("---")

    missing_coords = (df_points['Latitude'].isna().sum() +
                      df_points['Longitude'].isna().sum() +
                      df_points['zero_Latitude'].isna().sum() +
                      df_points['zero_Longitude'].isna().sum())

    if missing_coords > 0:
        st.warning(f"⚠️ Є {missing_coords} пропущених координат. Визначте їх на карті або в таблиці.")

    col_run, col_info = st.columns([1, 2])
    with col_run:
        run_clicked = st.button("🚀 Запустити оптимізацію", type="primary",
                                use_container_width=True)
    with col_info:
        st.caption(f"TOLERANCE={params['tolerance']}, MAX_STOPS={params['max_stops']}, "
                   f"N_SAMPLES={params['n_samples']}, SEED={params['seed']}")

    if run_clicked:
        log_messages = []
        log_container = st.empty()
        progress_bar = st.progress(0, text="Ініціалізація...")

        def log_callback(msg):
            log_messages.append(msg)
            log_container.markdown(
                '<div class="log-box">' + "<br>".join(log_messages[-40:]) + '</div>',
                unsafe_allow_html=True)

        with st.spinner("Оптимізація маршрутів..."):
            start_time = time.time()
            progress_bar.progress(10, text="Завантаження відстаней...")

            xlsx_bytes, all_rows, summary_rows, odometr_map = run_optimization(
                st.session_state.df_points, df_dates, df_plan, params,
                log_callback=log_callback)

            elapsed = time.time() - start_time
            progress_bar.progress(100, text=f"Готово за {elapsed:.0f} сек!")

        if xlsx_bytes is not None:
            st.session_state.result_xlsx = xlsx_bytes
            st.session_state.summary_rows = summary_rows
            st.session_state.all_rows = all_rows
            st.session_state.odometr_map = odometr_map
            st.success(f"✅ Оптимізацію завершено за {elapsed:.1f} сек!")

    # ── Results display ──
    if 'result_xlsx' in st.session_state and st.session_state.result_xlsx is not None:
        st.markdown("---")
        st.markdown("### 📋 Результати")

        summary = st.session_state.summary_rows
        odometr_map = st.session_state.get('odometr_map', {})

        if summary:
            tolerance_pct = params['tolerance'] * 100

            sum_data = []
            for s in summary:
                ok = abs(s["dev_pct"]) <= tolerance_pct
                odo_val = odometr_map.get((s["track"], s["period"]), "")
                odo_str = f"{odo_val:,.1f}" if isinstance(odo_val, (int, float)) else "—"
                sum_data.append({
                    "Трек": s["track"],
                    "Період": s["period"],
                    "План, км": f'{s["plan"]:.0f}',
                    "Факт, км": f'{s["fact"]:.1f}',
                    "Відхилення": f'{s["dev_pct"]:+.1f}%',
                    "Статус": "✓" if ok else "✗",
                    "Одометр": odo_str,
                })

            st.dataframe(pd.DataFrame(sum_data), use_container_width=True)

            ok_cnt = sum(1 for s in summary if abs(s["dev_pct"]) <= tolerance_pct)
            total_cnt = len(summary)

            rcols = st.columns(3)
            with rcols[0]:
                st.markdown(f'<div class="stat-card"><h3>{total_cnt}</h3><p>Періодів</p></div>',
                            unsafe_allow_html=True)
            with rcols[1]:
                st.markdown(f'<div class="stat-card"><h3 style="color:#2e7d32">{ok_cnt} ✓</h3>'
                            f'<p>В межах ±{tolerance_pct:.0f}%</p></div>',
                            unsafe_allow_html=True)
            with rcols[2]:
                fail_cnt = total_cnt - ok_cnt
                color = "#c62828" if fail_cnt > 0 else "#2e7d32"
                st.markdown(f'<div class="stat-card"><h3 style="color:{color}">{fail_cnt} ✗</h3>'
                            f'<p>Поза межами</p></div>',
                            unsafe_allow_html=True)

        st.markdown("")
        st.download_button(
            label="📥 Завантажити v7_optimized.xlsx",
            data=st.session_state.result_xlsx,
            file_name="v7_optimized.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )


if __name__ == "__main__":
    main()
